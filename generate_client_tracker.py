#!/usr/bin/env python3
"""
Lead Management Tracker Generator for Leads to Profit
Creates a comprehensive Excel file for daily client/lead management
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_lead_tracker():
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ========== SHEET 1: Daily Lead Tracker ==========
    ws_leads = wb.create_sheet("Daily Lead Tracker")

    # Headers for lead tracking
    headers = [
        "Lead ID", "Date Submitted", "Name", "Email", "Phone",
        "Service Interest", "Lead Source", "Lead Status", "Lead Score",
        "First Contact Date", "Response Time (hrs)", "Last Follow-Up",
        "Next Follow-Up", "Estimated Value", "Assigned To",
        "Priority", "Notes", "Outcome"
    ]

    # Style headers
    header_fill = PatternFill(start_color="00A651", end_color="00A651", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws_leads.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Add sample data
    sample_data = [
        [1, datetime.now().strftime("%Y-%m-%d"), "John Smith", "john@example.com", "555-0123",
         "Meta Ads + Google Ads", "Website Form", "Contacted", 85,
         datetime.now().strftime("%Y-%m-%d"), 2, datetime.now().strftime("%Y-%m-%d"),
         (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d"), "$5,000", "Your Name",
         "High", "Interested in scaling e-commerce store", ""],

        [2, datetime.now().strftime("%Y-%m-%d"), "Sarah Johnson", "sarah@company.com", "555-0456",
         "Lead Generation", "Referral", "Qualified", 90,
         datetime.now().strftime("%Y-%m-%d"), 1, datetime.now().strftime("%Y-%m-%d"),
         (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d"), "$8,000", "Your Name",
         "High", "B2B SaaS company, ready to start", ""],

        [3, (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d"), "Mike Davis", "mike@business.com", "",
         "CRO", "Website Form", "New", 65,
         "", "", "",
         datetime.now().strftime("%Y-%m-%d"), "$3,000", "",
         "Medium", "Left message, awaiting callback", ""],
    ]

    for row_num, data in enumerate(sample_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws_leads.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            if col_num in [2, 10, 12, 13]:  # Date columns
                cell.alignment = Alignment(horizontal='center')
            elif col_num == 9:  # Lead score
                cell.alignment = Alignment(horizontal='center')
                if isinstance(value, int):
                    if value >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Set column widths
    column_widths = [8, 12, 15, 20, 12, 20, 15, 12, 10, 15, 15, 12, 12, 12, 12, 10, 30, 12]
    for i, width in enumerate(column_widths, 1):
        ws_leads.column_dimensions[get_column_letter(i)].width = width

    # Freeze first row
    ws_leads.freeze_panes = 'A2'

    # ========== SHEET 2: KPI Dashboard ==========
    ws_kpi = wb.create_sheet("KPI Dashboard")

    # Title
    ws_kpi['A1'] = 'LEADS TO PROFIT - KPI DASHBOARD'
    ws_kpi['A1'].font = Font(bold=True, size=16, color="00A651")
    ws_kpi.merge_cells('A1:D1')

    ws_kpi['A2'] = f'Month: {datetime.now().strftime("%B %Y")}'
    ws_kpi['A2'].font = Font(size=12)
    ws_kpi.merge_cells('A2:D2')

    # KPI Headers
    kpi_row = 4
    ws_kpi[f'A{kpi_row}'] = 'KPI Metric'
    ws_kpi[f'B{kpi_row}'] = 'Target'
    ws_kpi[f'C{kpi_row}'] = 'Actual'
    ws_kpi[f'D{kpi_row}'] = 'Status'

    for col in ['A', 'B', 'C', 'D']:
        ws_kpi[f'{col}{kpi_row}'].fill = header_fill
        ws_kpi[f'{col}{kpi_row}'].font = header_font
        ws_kpi[f'{col}{kpi_row}'].alignment = Alignment(horizontal='center')
        ws_kpi[f'{col}{kpi_row}'].border = thin_border

    # KPI data
    kpis = [
        ['Total Leads', 50, 42, '84%'],
        ['Lead Response Time (avg hrs)', 2, 1.5, '✓ On Track'],
        ['Qualified Leads', 25, 28, '112%'],
        ['Conversion Rate', '50%', '52%', '✓ Above Target'],
        ['Proposals Sent', 15, 12, '80%'],
        ['Deals Won', 8, 6, '75%'],
        ['Average Deal Value', '$5,000', '$5,500', '110%'],
        ['Total Revenue', '$40,000', '$33,000', '83%'],
        ['Cost Per Lead', '$50', '$45', '✓ Below Target'],
        ['Customer Acquisition Cost', '$500', '$475', '✓ Below Target'],
        ['Lead-to-Customer Rate', '20%', '14%', '70%'],
        ['Monthly Recurring Revenue', '$15,000', '$12,500', '83%'],
    ]

    for idx, kpi in enumerate(kpis, kpi_row + 1):
        for col_num, value in enumerate(kpi, 1):
            cell = ws_kpi.cell(row=idx, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_num > 1 else 'left')

            # Status column coloring
            if col_num == 4:
                if '✓' in str(value) or (isinstance(value, str) and '%' in value and
                    value.replace('%', '').isdigit() and int(value.replace('%', '')) >= 100):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif isinstance(value, str) and '%' in value:
                    try:
                        percent = int(value.replace('%', ''))
                        if percent < 80:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif percent < 100:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    except:
                        pass

    ws_kpi.column_dimensions['A'].width = 30
    ws_kpi.column_dimensions['B'].width = 15
    ws_kpi.column_dimensions['C'].width = 15
    ws_kpi.column_dimensions['D'].width = 15

    # ========== SHEET 3: Sales Pipeline ==========
    ws_pipeline = wb.create_sheet("Sales Pipeline")

    pipeline_headers = [
        "Lead ID", "Name", "Company", "Service Interest", "Stage",
        "Days in Stage", "Deal Value", "Probability", "Expected Close",
        "Last Activity", "Next Steps"
    ]

    for col_num, header in enumerate(pipeline_headers, 1):
        cell = ws_pipeline.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Pipeline stages data
    pipeline_data = [
        [1, "John Smith", "Smith E-commerce", "Meta Ads + Google", "Proposal Sent", 3, "$5,000", "60%",
         (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"), "Sent proposal", "Follow up on Tuesday"],

        [2, "Sarah Johnson", "TechStart SaaS", "Lead Generation", "Negotiation", 5, "$8,000", "80%",
         (datetime.now() + timedelta(days=5)).strftime("%Y-%m-%d"), "Price discussion", "Send contract"],

        [3, "Mike Davis", "Davis Consulting", "CRO", "Discovery", 2, "$3,000", "30%",
         (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"), "Initial call", "Schedule audit"],

        [4, "Lisa Chen", "Chen Marketing", "Email Marketing", "Qualified", 1, "$4,500", "40%",
         (datetime.now() + timedelta(days=10)).strftime("%Y-%m-%d"), "Needs assessment", "Send case studies"],
    ]

    for row_num, data in enumerate(pipeline_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws_pipeline.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            if col_num in [6, 8]:
                cell.alignment = Alignment(horizontal='center')
            elif col_num in [7]:
                cell.alignment = Alignment(horizontal='right')

            # Color code by stage
            if col_num == 5:
                stage = value
                if stage == "Negotiation":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif stage == "Proposal Sent":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    pipeline_widths = [8, 15, 18, 18, 15, 12, 12, 12, 12, 15, 25]
    for i, width in enumerate(pipeline_widths, 1):
        ws_pipeline.column_dimensions[get_column_letter(i)].width = width

    ws_pipeline.freeze_panes = 'A2'

    # ========== SHEET 4: Weekly Activity Log ==========
    ws_activity = wb.create_sheet("Weekly Activity Log")

    activity_headers = [
        "Date", "Time", "Lead Name", "Activity Type", "Channel",
        "Duration", "Outcome", "Follow-Up Required", "Notes"
    ]

    for col_num, header in enumerate(activity_headers, 1):
        cell = ws_activity.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Sample activity data
    activity_data = [
        [datetime.now().strftime("%Y-%m-%d"), "10:30 AM", "John Smith", "Discovery Call", "Phone",
         "30 min", "Positive - Ready for proposal", "Yes - Send proposal", "Discussed budget and timeline"],

        [datetime.now().strftime("%Y-%m-%d"), "2:00 PM", "Sarah Johnson", "Follow-Up Email", "Email",
         "5 min", "Sent contract", "Yes - Call Friday", "Addressed pricing questions"],

        [(datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d"), "11:00 AM", "Mike Davis", "Initial Contact", "Phone",
         "15 min", "Left voicemail", "Yes - Call tomorrow", "No answer, try mobile"],
    ]

    for row_num, data in enumerate(activity_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws_activity.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border

    activity_widths = [12, 10, 15, 15, 10, 10, 20, 15, 30]
    for i, width in enumerate(activity_widths, 1):
        ws_activity.column_dimensions[get_column_letter(i)].width = width

    ws_activity.freeze_panes = 'A2'

    # ========== SHEET 5: Data Validation & Instructions ==========
    ws_guide = wb.create_sheet("Setup Guide")

    guide_content = [
        ["LEADS TO PROFIT - CLIENT MANAGEMENT TRACKER", ""],
        ["", ""],
        ["HOW TO USE THIS TRACKER:", ""],
        ["", ""],
        ["1. Daily Lead Tracker", "Record all new leads and update their status daily"],
        ["", "- Lead Status options: New, Contacted, Qualified, Proposal Sent, Won, Lost"],
        ["", "- Lead Score: Rate 0-100 based on fit and interest"],
        ["", "- Priority: High, Medium, Low"],
        ["", ""],
        ["2. KPI Dashboard", "Track monthly performance metrics"],
        ["", "- Update 'Actual' column with real numbers"],
        ["", "- Green = on/above target, Yellow = close, Red = below"],
        ["", ""],
        ["3. Sales Pipeline", "Manage deals in progress"],
        ["", "- Stages: Discovery → Qualified → Proposal → Negotiation → Won/Lost"],
        ["", "- Update 'Days in Stage' to track velocity"],
        ["", ""],
        ["4. Weekly Activity Log", "Record all client interactions"],
        ["", "- Log calls, emails, meetings daily"],
        ["", "- Track follow-up requirements"],
        ["", ""],
        ["BEST PRACTICES:", ""],
        ["", ""],
        ["✓ Response Time", "Contact new leads within 2 hours (3x better conversion)"],
        ["✓ Lead Scoring", "Focus on leads 70+ score first"],
        ["✓ Follow-Up", "Set next follow-up date for every lead"],
        ["✓ Pipeline Review", "Review pipeline weekly to identify stuck deals"],
        ["✓ Data Entry", "Update tracker daily for accuracy"],
        ["", ""],
        ["SERVICE CATEGORIES:", ""],
        ["", "Lead Generation, Meta Ads, Google Ads, CRO, Email Marketing, Strategy"],
        ["", ""],
        ["LEAD SOURCES:", ""],
        ["", "Website Form, Referral, LinkedIn, Cold Outreach, Networking Event, Other"],
    ]

    for row_num, (label, value) in enumerate(guide_content, 1):
        ws_guide[f'A{row_num}'] = label
        ws_guide[f'B{row_num}'] = value

        if row_num == 1:
            ws_guide[f'A{row_num}'].font = Font(bold=True, size=14, color="00A651")
            ws_guide.merge_cells(f'A{row_num}:B{row_num}')
        elif "HOW TO USE" in label or "BEST PRACTICES" in label or "SERVICE" in label or "LEAD SOURCES" in label:
            ws_guide[f'A{row_num}'].font = Font(bold=True, size=12, color="00A651")
        elif label.startswith(('1.', '2.', '3.', '4.', '✓')):
            ws_guide[f'A{row_num}'].font = Font(bold=True)

    ws_guide.column_dimensions['A'].width = 25
    ws_guide.column_dimensions['B'].width = 60

    # Save the workbook
    filename = f"Leads_to_Profit_Client_Tracker_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✓ Successfully created: {filename}")
    return filename

if __name__ == "__main__":
    create_lead_tracker()
